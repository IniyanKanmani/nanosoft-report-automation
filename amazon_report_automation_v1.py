import pandas as pd
import numpy as np
import openpyxl as el

data = pd.read_csv('input_file.csv', header=7)

adjustment = data.loc[data['type'] == 'Adjustment']['total'].astype(np.float64).sum()

fba_inventory_fee = data.loc[data['type'] == 'FBA Inventory Fee']['total'].astype(np.float64).sum()* -1

service_fee = data.loc[data['type'] == 'Service Fee']['total'].astype(np.float64).sum() * -1

product_sales_tax = data.loc[data['type'] == 'Order']['product sales tax'].astype(np.float64).sum()

wb = el.load_workbook('output report.xlsx')

ws = wb.active

ws['B16'].value = adjustment

ws['B9'].value = fba_inventory_fee

ws['B8'].value = service_fee

ws['B5'].value = product_sales_tax

wb.save('output report.xlsx')

