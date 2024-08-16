import pandas as pd
import numpy as np
import openpyxl as el
import customtkinter

customtkinter.set_appearance_mode('System')
customtkinter.set_default_color_theme("blue")

root = customtkinter.CTk()
root.geometry("1000x725")

language = ''

input_file = ''
input_filename = customtkinter.StringVar()
# input_filename.set('Choose Input File' if language == 'english' else 'Elija el archivo de entrada')

output_file = ''
output_filename = customtkinter.StringVar()
# output_filename.set('Choose Output File' if language == 'english' else 'Elija el archivo de salida')

def get_input_filename():
    
    global input_file, input_filename
    input_file = customtkinter.filedialog.askopenfilename()
    input_filename.set(input_file.split('/')[-1])


def get_output_filename():
    
    global output_file, output_filename
    output_file = customtkinter.filedialog.askopenfilename()
    output_filename.set(output_file.split('/')[-1])


def process_files():
    
    global root, input_file, output_file

    data = pd.read_excel(input_file, header=7)

    # data.head()

    adjustment = data.loc[data['type'] == 'Adjustment']['total'].astype(np.float64).sum()

    fba_inventory_fee = data.loc[data['type'] == 'FBA Inventory Fee']['total'].astype(np.float64).sum()* -1

    service_fee = data.loc[data['type'] == 'Service Fee']['total'].astype(np.float64).sum() * -1

    product_sales_tax = data.loc[data['type'] == 'Order']['product sales tax'].astype(np.float64).sum()

    paid_transaction_gross = data.loc[data['type'] == 'Order']['product sales'].astype(np.float64).sum() + data.loc[data['type'] == 'Order']['product sales tax'].astype(np.float64).sum()

    sales_after_amazon_sales_fees = data.loc[data['type'] == 'Order']['total'].astype(np.float64).sum()

    amazon_fees_on_orders = paid_transaction_gross - sales_after_amazon_sales_fees

    returns = data.loc[data['type'] == 'Refund']['total'].astype(np.float64).sum() * -1

    reclaimed_vat = data.loc[data['type'] == 'Refund']['product sales tax'].astype(np.float64).sum() * -1

    # items_returned_to_stock = data.loc[data['type'] == 'Refund']['product sales'].astype(np.float64).sum() * -1

    wb = el.load_workbook(output_file)

    ws = wb.active

    ws['B16'].value = adjustment

    ws['B9'].value = fba_inventory_fee

    ws['B8'].value = service_fee

    ws['B5'].value = product_sales_tax
    
    ws['B3'].value = paid_transaction_gross
    
    ws['B4'].value = sales_after_amazon_sales_fees
    
    ws['B6'].value = amazon_fees_on_orders

    ws['B13'].value = returns

    ws['B14'].value = reclaimed_vat

    # ws['B15'].value = items_returned_to_stock

    wb.save(output_file)
    
    root.destroy()


def switch_to_second_page():
    
    global primary_frame
    
    primary_frame.pack_forget()

    secondary_frame = customtkinter.CTkFrame(master=root)
    secondary_frame.pack(pady=30, padx=60, fill='both', expand=True)

    input_frame = customtkinter.CTkFrame(master=secondary_frame)
    input_frame.pack(pady=(40, 25), padx=60, fill='both')

    input_label_frame = customtkinter.CTkFrame(master=input_frame)
    input_label_frame.pack(pady=(40, 20), padx=60, fill='both')

    input_filename.set('Choose Input File' if language == 'english' else 'Elija el archivo de entrada')
    input_label = customtkinter.CTkLabel(master=input_label_frame, textvariable=input_filename)
    input_label.pack(pady=15, padx=10)

    input_button = customtkinter.CTkButton(master=input_frame, text='Import' if language == 'english' else 'Importar', command=get_input_filename)
    input_button.pack(pady=(20,40), padx=10)

    output_frame = customtkinter.CTkFrame(master=secondary_frame)
    output_frame.pack(pady=(25, 20), padx=60, fill='both')

    output_label_frame = customtkinter.CTkFrame(master=output_frame)
    output_label_frame.pack(pady=(40, 20), padx=60, fill='both')

    output_filename.set('Choose Output File' if language == 'english' else 'Elija el archivo de salida')
    output_label = customtkinter.CTkLabel(master=output_label_frame, textvariable=output_filename)
    output_label.pack(pady=15, padx=10)

    output_button = customtkinter.CTkButton(master=output_frame, text='Import' if language == 'english' else 'Importar', command=get_output_filename)
    output_button.pack(pady=(20, 40), padx=10)

    process_button = customtkinter.CTkButton(master=secondary_frame, text='Process Files' if language == 'english' else 'Archivos de proceso', command=process_files)
    process_button.pack(pady=(25,30), padx=10)


def choose_language(lang):
    global language
    language = lang
    switch_to_second_page()


primary_frame = customtkinter.CTkFrame(master=root)
primary_frame.pack(pady=30, padx=60, fill='both', expand=True)

choose_language_label = customtkinter.CTkLabel(master=primary_frame, text="Hola Luke, Which language do you prefer today?")
choose_language_label.pack(pady=(150, 100), padx=10)

english_button = customtkinter.CTkButton(master=primary_frame, text='English', command=lambda: choose_language('english'))
english_button.pack(pady=(20,40), padx=10)

spanish_button = customtkinter.CTkButton(master=primary_frame, text='Español', command=lambda: choose_language('español'))
spanish_button.pack(pady=(20,40), padx=10)

root.mainloop()
