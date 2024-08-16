import pandas as pd
import numpy as np
import openpyxl as el

data = pd.read_excel('./Input file.xlsx', header=7)

# data.head()

adjustment = data.loc[data['type'] == 'Adjustment']['total'].astype(np.float64).sum()

fba_inventory_fee = data.loc[data['type'] == 'FBA Inventory Fee']['total'].astype(np.float64).sum()* -1

service_fee = data.loc[data['type'] == 'Service Fee']['total'].astype(np.float64).sum() * -1

product_sales_tax = data.loc[data['type'] == 'Order']['product sales tax'].astype(np.float64).sum()

paid_transaction_gross = data.loc[data['type'] == 'Order']['product sales'].astype(np.float64).sum() + data.loc[data['type'] == 'Order']['product sales tax'].astype(np.float64).sum()

sales_after_amazon_sales_fees = data.loc[data['type'] == 'Order']['total'].astype(np.float64).sum()

amazon_fees_on_orders = paid_transaction_gross - sales_after_amazon_sales_fees

returns = data.loc[data['type'] == 'Refund']['total'].astype(np.float64).sum() * -1

reclaimed_vat = data.loc[data['type'] == 'Refund']['product sales tax'].astype(np.float64).sum() * -1

# items_returned_to_stock = data.loc[data['type'] == 'Refund']['product sales'].astype(np.float64).sum() * -1

order_sku_quantity = data.loc[data['type'] == 'Order'][['sku', 'quantity']]
unique_order_sku = order_sku_quantity['sku'].unique()
order_sku_list = []
order_quantity_list = []

for sku in unique_order_sku:
    order_sku_list.append(sku)
    order_quantity_list.append(order_sku_quantity.loc[data['sku'] == sku]['quantity'].sum())
    
new_order_sku_quantity = pd.DataFrame({'sku': order_sku_list, 'quantity': order_quantity_list})

refund_sku_quantity = data.loc[data['type'] == 'Refund'][['sku', 'quantity']]
unique_refund_sku = refund_sku_quantity['sku'].unique()
refund_sku_list = []
refund_quantity_list = []

for sku in unique_refund_sku:
    refund_sku_list.append(sku)
    refund_quantity_list.append(refund_sku_quantity.loc[data['sku'] == sku]['quantity'].sum())
    
new_refund_sku_quantity = pd.DataFrame({'sku': refund_sku_list, 'quantity': refund_quantity_list})

wb = el.load_workbook('./output report.xlsx')

ws = wb.active

ws['B16'].value = adjustment

ws['B9'].value = fba_inventory_fee

ws['B8'].value = service_fee

ws['B5'].value = product_sales_tax

ws['B3'].value = paid_transaction_gross

ws['B4'].value = sales_after_amazon_sales_fees

ws['B6'].value = amazon_fees_on_orders

ws['B13'].value = returns

ws['B14'].value = reclaimed_vat

# ws['B15'].value = items_returned_to_stock

wb.save('./output report.xlsx')

i = 27

while ws.cell(row=i, column=1).value != None:
    if ws.cell(row=i, column=1).value in unique_order_sku:
        ws.cell(row=i, column=5).value = new_order_sku_quantity.loc[new_order_sku_quantity['sku'] == ws.cell(row=i, column=1).value]['quantity'].item()
    i = i + 1
    
wb.save('./output report.xlsx')

i = 27

while ws.cell(row=i, column=1).value != None:
    if ws.cell(row=i, column=1).value in unique_refund_sku:
        ws.cell(row=i, column=14).value = new_refund_sku_quantity.loc[new_refund_sku_quantity['sku'] == ws.cell(row=i, column=1).value]['quantity'].item()
    i = i + 1

wb.save('./output report.xlsx')

i = 27

while ws.cell(row=i, column=1).value != None:
    if ws.cell(row=i, column=1).value in unique_refund_sku:
        ws.cell(row=i, column=15).value = ws.cell(row=i, column=4).value * ws.cell(row=i, column=14).value
    i = i + 1

wb.save('./output report.xlsx')
