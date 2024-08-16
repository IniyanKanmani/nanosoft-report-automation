import pandas as pd
import numpy as np
import openpyxl as el
import customtkinter
import time

customtkinter.set_appearance_mode('System')
customtkinter.set_default_color_theme("blue")

root = customtkinter.CTk()
root.title('Amazon Report Generator')
root.geometry("1000x725")

input_file = ''
input_filename = customtkinter.StringVar()

output_file = ''
output_filename = customtkinter.StringVar()

def get_input_filename():

    global input_file, input_filename
    input_file = customtkinter.filedialog.askopenfilename()
    input_filename.set(input_file.split('/')[-1])


def get_output_filename():

    global output_file, output_filename
    output_file = customtkinter.filedialog.askopenfilename()
    output_filename.set(output_file.split('/')[-1])


def switch_to_end_screen():

    global root, main_frame

    main_frame.pack_forget()

    end_frame = customtkinter.CTkFrame(master=root)
    end_frame.pack(pady=30, padx=60, fill='both', expand=True)

    result_label = customtkinter.CTkLabel(master=end_frame, text="Report is generated Successfully", )
    result_label.place(relx=0.5, rely=0.5, anchor='center')
    
    root.after(2000, lambda: root.destroy())
    

def switch_to_error_end_screen(e):

    global root, main_frame

    main_frame.pack_forget()

    error_end_frame = customtkinter.CTkFrame(master=root)
    error_end_frame.pack(pady=30, padx=60, fill='both', expand=True)

    result_label = customtkinter.CTkLabel(master=error_end_frame, text=f"Error: {e}", )
    result_label.place(relx=0.5, rely=0.5, anchor='center')


def convert_to_numeric(data, column_name):
    if data[column_name].dtype != np.float64 and data[column_name].dtype != np.int64:
        for i in range(len(data)):
            data[column_name][i] = float(data[column_name][i].replace(',', ''))
        data[column_name] = pd.to_numeric(data[column_name])
    return data


def process_files():
    
    try:

        global root, input_file, output_file

        data = pd.read_csv(input_file, header=7)

        data = convert_to_numeric(data, 'quantity')
        data = convert_to_numeric(data, 'product sales')
        data = convert_to_numeric(data, 'product sales tax')
        data = convert_to_numeric(data, 'total')

        adjustment = data.loc[data['type'] == 'Adjustment']['total'].astype(np.float64).sum()

        fba_inventory_fee = data.loc[data['type'] == 'FBA Inventory Fee']['total'].astype(np.float64).sum() * -1

        service_fee = data.loc[data['type'] == 'Service Fee']['total'].astype(np.float64).sum() * -1

        product_sales_tax = data.loc[data['type'] == 'Order']['product sales tax'].astype(np.float64).sum()

        paid_transaction_gross = data.loc[data['type'] == 'Order']['product sales'].astype(np.float64).sum() + data.loc[data['type'] == 'Order']['product sales tax'].astype(np.float64).sum()

        sales_after_amazon_sales_fees = data.loc[data['type'] == 'Order']['total'].astype(np.float64).sum()

        amazon_fees_on_orders = paid_transaction_gross - sales_after_amazon_sales_fees

        returns = data.loc[data['type'] == 'Refund']['total'].astype(np.float64).sum() * -1

        reclaimed_vat = data.loc[data['type'] == 'Refund']['product sales tax'].astype(np.float64).sum() * -1

        retrocharge = data.loc[data['type'] == 'Retrocharge']['total'].astype(np.float64).sum()

        order_sku_quantity_tax_total = data.loc[data['type'] == 'Order'][['sku', 'quantity', 'product sales tax', 'total']]
        order_sku_quantity_tax_total.head()
        new_order_sku_quantity_tax_total = order_sku_quantity_tax_total.groupby(['sku'], as_index=False).sum()

        refund_sku_quantity = data.loc[data['type'] == 'Refund'][['sku', 'quantity']]
        new_refund_sku_quantity = refund_sku_quantity.groupby(['sku'], as_index=False).sum()

        order_refund_sku = np.concatenate((data.loc[data['type'] == 'Order']['sku'].unique(), data.loc[data['type'] == 'Refund']['sku'].unique()))
        order_refund_sku = pd.Series(order_refund_sku).dropna().to_numpy()
        order_refund_sku = np.unique(order_refund_sku)

        wb = el.load_workbook(output_file)

        ws1 = wb.active

        ws1['B16'].value = adjustment

        ws1['B9'].value = fba_inventory_fee

        ws1['B8'].value = service_fee

        ws1['B5'].value = product_sales_tax

        ws1['B3'].value = paid_transaction_gross

        ws1['B4'].value = sales_after_amazon_sales_fees

        ws1['B6'].value = amazon_fees_on_orders

        ws1['B13'].value = returns

        ws1['B14'].value = reclaimed_vat
        
        ws1['B21'].value = retrocharge

        wb.save(output_file)

        i = 27

        while ws1.cell(row=i, column=1).value != None:
            if ws1.cell(row=i, column=1).value in new_order_sku_quantity_tax_total['sku'].unique():
                ws1.cell(row=i, column=5).value = new_order_sku_quantity_tax_total.loc[new_order_sku_quantity_tax_total['sku'] == ws1.cell(row=i, column=1).value]['quantity'].astype(np.float64).item()
            i = i + 1

        wb.save(output_file)

        i = 27

        while ws1.cell(row=i, column=1).value != None:
            if ws1.cell(row=i, column=1).value in new_refund_sku_quantity['sku'].unique():
                ws1.cell(row=i, column=14).value = new_refund_sku_quantity.loc[new_refund_sku_quantity['sku'] == ws1.cell(row=i, column=1).value]['quantity'].astype(np.float64).item()
            i = i + 1

        wb.save(output_file)

        i = 27

        while ws1.cell(row=i, column=1).value != None:
            if ws1.cell(row=i, column=1).value in new_refund_sku_quantity['sku'].unique():
                ws1.cell(row=i, column=15).value = ws1.cell(row=i, column=4).value * ws1.cell(row=i, column=14).value
            i = i + 1

        wb.save(output_file)

        i = 27

        while ws1.cell(row=i, column=1).value != None:
            if ws1.cell(row=i, column=1).value in new_order_sku_quantity_tax_total['sku'].unique():
                ws1.cell(row=i, column=6).value = new_order_sku_quantity_tax_total.loc[new_order_sku_quantity_tax_total['sku'] == ws1.cell(row=i, column=1).value]['product sales tax'].astype(np.float64).item()
            i = i + 1

        wb.save(output_file)

        i = 27

        while ws1.cell(row=i, column=1).value != None:
            if ws1.cell(row=i, column=1).value in new_order_sku_quantity_tax_total['sku'].unique():
                ws1.cell(row=i, column=7).value = new_order_sku_quantity_tax_total.loc[new_order_sku_quantity_tax_total['sku'] == ws1.cell(row=i, column=1).value]['total'].astype(np.float64).item()
            i = i + 1

        wb.save(output_file)

        report_file_sku = []

        i = 27

        while ws1.cell(row=i, column=1).value != None:
            report_file_sku.append(ws1.cell(row=i, column=1).value)
            i = i + 1

        ws2 = wb['Sheet2']

        for val in order_refund_sku:
            if val not in report_file_sku:
                i = 1
                while ws2.cell(row=i, column=1).value != None: 
                    i = i + 1
                ws2.cell(row=i, column=1).value = val

        wb.save(output_file)

        switch_to_end_screen()

    except Exception as e:
        
        switch_to_error_end_screen(e)

main_frame = customtkinter.CTkFrame(master=root)
main_frame.pack(pady=30, padx=60, fill='both', expand=True)

input_frame = customtkinter.CTkFrame(master=main_frame)
input_frame.pack(pady=(40, 25), padx=60, fill='both')

input_label_frame = customtkinter.CTkFrame(master=input_frame)
input_label_frame.pack(pady=(40, 20), padx=60, fill='both')

input_filename.set('Choose Input File')
input_label = customtkinter.CTkLabel(master=input_label_frame, textvariable=input_filename)
input_label.pack(pady=15, padx=10)

input_button = customtkinter.CTkButton(master=input_frame, text='Import', command=get_input_filename)
input_button.pack(pady=(20,40), padx=10)

output_frame = customtkinter.CTkFrame(master=main_frame)
output_frame.pack(pady=(25, 20), padx=60, fill='both')

output_label_frame = customtkinter.CTkFrame(master=output_frame)
output_label_frame.pack(pady=(40, 20), padx=60, fill='both')

output_filename.set('Choose Output File')
output_label = customtkinter.CTkLabel(master=output_label_frame, textvariable=output_filename)
output_label.pack(pady=15, padx=10)

output_button = customtkinter.CTkButton(master=output_frame, text='Import', command=get_output_filename)
output_button.pack(pady=(20, 40), padx=10)

process_button = customtkinter.CTkButton(master=main_frame, text='Process Files', command=process_files)
process_button.pack(pady=(25,30), padx=10)

root.mainloop()
